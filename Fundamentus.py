from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from time import sleep
import openpyxl
import os



class Fundamentus:

    def iniciar(self):
        self.raspagem_de_dados()
        self.criar_planilha()

    def raspagem_de_dados(self):
        chrome_options = Options()
        chrome_options.add_experimental_option(
            'excludeSwitches', ['enable-logging'])
        chrome_options.add_argument('--lang=pr-BR')
        chrome_options.add_argument('--disable-notifications')
        self.driver = webdriver.Chrome(executable_path=os.getcwd(
        ) + os.sep + 'chromedriver.exe', options=chrome_options)
        self.driver.set_window_size(800, 700)
        self.link = 'https://fundamentus.com.br/resultado.php'
        print(self.driver.title)
        self.lista_nome_papel = []
        self.lista_preco_ev = []
        self.driver.get(self.link)
        sleep(2)
        item = 1
        for i in range(36):
            lista_nomes = self.driver.find_elements_by_xpath(
            f'//span[@class="tips"]/a[{item}]')
            self.lista_nome_papel.append(lista_nomes[0].text)
            sleep(1)
            lista_precos = self.driver.find_elements_by_xpath(
            f'/html/body/div[1]/div[2]/table/tbody/tr[1]/td[{item}]')
            self.lista_preco_ev.append(lista_precos[0].text)
            item += 1
            sleep(1)
            print(f'\u001b[32m{"Escaneamento Concluido"}\u001b[0m')
            self.driver.quit()

    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        celulares = planilha['Sheet']
        celulares.title = 'Fundamentus'
        celulares['A1'] = 'Papel'
        celulares['B1'] = 'EV/EBIT'
        for nome, preco in zip(self.lista_nome_papel, self.lista_preco_ev):
            celulares.cell(column=1, row=index, value=nome)
            celulares.cell(column=2, row=index, value=preco)
            index += 1
        planilha.save("planilha_de_preços.xlsx")

        print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')


start = Fundamentus()
start.iniciar()
